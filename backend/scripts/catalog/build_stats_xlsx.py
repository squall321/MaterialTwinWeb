# 카탈로그 전체 통계를 xlsx 한 권으로 — 출처·물성·재료·해석 가능성을 시트별로 정리한다.
#
# 사용: build_stats_xlsx.py [출력경로]
#
# 시트 구성:
#   01 요약            전체 규모와 신뢰등급 분포
#   02 출처통계        출처 종류·발행처·DOI 보유·물성값 기여 상위
#   03 물성정의목록    128개 정의 전체 + 보유 건수·재료수·등급분포
#   04 도메인별통계    11개 도메인 집계
#   05 재료목록        425종 전체 + 물성수·등급·해석 가능 여부
#   06 해석준비도      해석 7종별 완비율과 병목
#   07 재료물성매트릭스 카테고리 × 주요 물성 커버리지
#   08 데이터품질      가정값·조건 보유·이상치 점검 결과
import pathlib
import sqlite3
import statistics as st
import sys
from collections import Counter, defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DB = '/home/koopark/claude/HEAXHub/var/app_data/materialtwin_web/materialtwin.db'
OUT = sys.argv[1] if len(sys.argv) > 1 else '/home/koopark/claude/MaterialTwinWeb/docs/카탈로그_통계.xlsx'

HDR_FILL = PatternFill("solid", fgColor="1F3864")
HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=13, color="1F3864")
SUB_FONT = Font(italic=True, size=9, color="595959")
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def sheet(wb, name, title, note=""):
    ws = wb.create_sheet(name)
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    if note:
        ws["A2"] = note
        ws["A2"].font = SUB_FONT
    ws.freeze_panes = "A5"
    return ws


def head(ws, row, cols):
    for j, c in enumerate(cols, 1):
        cell = ws.cell(row=row, column=j, value=c)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def body(ws, row0, rows, widths=None, numfmt=None):
    for i, r in enumerate(rows):
        for j, v in enumerate(r, 1):
            cell = ws.cell(row=row0 + i, column=j, value=v)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=isinstance(v, str) and len(str(v)) > 40)
            if numfmt and j in numfmt:
                cell.number_format = numfmt[j]
    for j, w in enumerate(widths or [], 1):
        ws.column_dimensions[get_column_letter(j)].width = w


def main():
    c = sqlite3.connect(DB)
    q1 = lambda s, *a: c.execute(s, a).fetchone()[0]
    wb = Workbook()
    wb.remove(wb.active)

    n_mat = q1("select count(*) from material")
    n_pv = q1("select count(*) from property_value")
    n_src = q1("select count(*) from source")
    n_def = q1("select count(*) from property_definition")
    n_used = q1("select count(distinct property_key) from property_value")

    # ── 01 요약 ──────────────────────────────────────────────────────────────
    ws = sheet(wb, "01_요약", "MaterialTwin 카탈로그 통계 요약",
               "물성값 1건 = (재료 × 물성 × 조건 × 출처) 하나. 같은 물성도 조건·출처가 다르면 별개 행이다.")
    head(ws, 4, ["항목", "값", "설명"])
    tiers = dict(c.execute("select quality_tier, count(*) from property_value group by 1").fetchall())
    meth = dict(c.execute("select method, count(*) from property_value group by 1").fetchall())
    rows = [
        ("재료", n_mat, "카탈로그에 등록된 재료 종수"),
        ("물성값", n_pv, "조건·출처별로 분리된 개별 값"),
        ("출처", n_src, "데이터시트·논문·핸드북 등 원문 단위"),
        ("물성 정의", n_def, f"정의된 물성 종류. 값이 있는 것 {n_used}종"),
        ("시편", q1("select count(*) from specimen"), "인장·완화 시험 시편"),
        ("시험", q1("select count(*) from test"), "등록된 시험 기록"),
        ("", "", ""),
        ("tier1 (실측)", tiers.get(1, 0), "그 재료·제품에 대해 문서에 인쇄된 시험값"),
        ("tier2 (핸드북)", tiers.get(2, 0), "핸드북·표준·인증 DB"),
        ("tier3 (2차·클래스)", tiers.get(3, 0), "2차 인용, 클래스 대표값, 그래프 디지타이즈"),
        ("tier4 (계산·가정)", tiers.get(4, 0), "계산 유도값, 가정값"),
        ("", "", ""),
        ("method=measured", meth.get("measured", 0), "측정값"),
        ("method=handbook", meth.get("handbook", 0), "핸드북·표준 인용"),
        ("method=computed", meth.get("computed", 0), "다른 물성에서 계산·피팅"),
        ("method=estimated", meth.get("estimated", 0), "추정·가정"),
        ("", "", ""),
        ("가정값(assumption 표지)", q1("select count(*) from property_value where instr(coalesce(conditions,''),'assumption')>0"),
         "클래스 대표를 빌려 온 값. UI에 '가정' 배지로 표시되고 실측이 들어오면 대체된다"),
        ("조건(conditions) 보유", q1("select count(*) from property_value where conditions is not null"),
         "온도·파장·방향·율속 등 측정 조건이 구조화된 값"),
        ("DOI 보유 출처", q1("select count(*) from source where coalesce(doi,'')<>''"), "논문 출처의 검증 가능성"),
        ("URL 보유 출처", q1("select count(*) from source where coalesce(url,'')<>''"), "원문 재확인 경로"),
    ]
    body(ws, 5, rows, [26, 12, 62], {2: "#,##0"})

    # ── 02 출처통계 ──────────────────────────────────────────────────────────
    ws = sheet(wb, "02_출처통계", "데이터가 어디서 왔는가",
               "kind는 원문의 종류다. datasheet=벤더 기술자료, journal=학술논문, database=2차 물성 DB.")
    head(ws, 4, ["출처 종류", "출처 수", "물성값 수", "물성값 비율", "DOI 보유", "URL 보유"])
    rows = []
    for kind, ns in c.execute("select kind, count(*) from source group by kind order by 2 desc"):
        npv = q1("select count(*) from property_value pv join source s on s.id=pv.source_id where s.kind=?", kind)
        ndoi = q1("select count(*) from source where kind=? and coalesce(doi,'')<>''", kind)
        nurl = q1("select count(*) from source where kind=? and coalesce(url,'')<>''", kind)
        rows.append((kind, ns, npv, npv / n_pv, ndoi, nurl))
    body(ws, 5, rows, [14, 10, 12, 12, 10, 10], {4: "0.0%"})

    r0 = 5 + len(rows) + 2
    ws.cell(row=r0, column=1, value="발행처 상위 30 (물성값 기여 순)").font = TITLE_FONT
    head(ws, r0 + 1, ["발행처", "출처 수", "물성값 수", "주 종류"])
    pub = c.execute("""select coalesce(nullif(trim(s.publisher),''),'(미기재)') p, count(distinct s.id),
        count(pv.id), s.kind from source s left join property_value pv on pv.source_id=s.id
        group by p order by count(pv.id) desc limit 30""").fetchall()
    body(ws, r0 + 2, pub, [46, 10, 12, 14])

    r1 = r0 + 2 + len(pub) + 2
    ws.cell(row=r1, column=1, value="개별 출처 상위 30 (물성값 기여 순)").font = TITLE_FONT
    head(ws, r1 + 1, ["출처 제목", "종류", "발행처", "연도", "DOI", "물성값 수"])
    top = c.execute("""select substr(s.title,1,90), s.kind, coalesce(s.publisher,''), coalesce(s.year,''),
        coalesce(s.doi,''), count(pv.id) n from source s join property_value pv on pv.source_id=s.id
        group by s.id order by n desc limit 30""").fetchall()
    body(ws, r1 + 2, top, [70, 12, 28, 8, 26, 10])

    # ── 03 물성정의목록 ──────────────────────────────────────────────────────
    ws = sheet(wb, "03_물성정의목록", "정의된 물성 전체 목록",
               "값이 0건인 정의는 아직 수집되지 않은 항목이다. si_unit은 저장 단위이며 모든 값이 이 단위로 정규화된다.")
    head(ws, 4, ["물성 키", "도메인", "이름", "SI 단위", "값 건수", "보유 재료수",
                 "tier1", "tier2", "tier3", "tier4", "최소", "중앙값", "최대"])
    rows = []
    for key, dom, nm, unit in c.execute("select key, domain, name, si_unit from property_definition order by domain, key"):
        vals = [v for (v,) in c.execute("select value_num from property_value where property_key=? and value_num is not null", (key,))]
        nrow = q1("select count(*) from property_value where property_key=?", key)
        nmat = q1("select count(distinct material_id) from property_value where property_key=?", key)
        tt = dict(c.execute("select quality_tier, count(*) from property_value where property_key=? group by 1", (key,)).fetchall())
        rows.append((key, dom, nm, unit or "", nrow, nmat,
                     tt.get(1, 0), tt.get(2, 0), tt.get(3, 0), tt.get(4, 0),
                     min(vals) if vals else None, st.median(vals) if vals else None, max(vals) if vals else None))
    body(ws, 5, rows, [38, 12, 26, 16, 9, 10, 7, 7, 7, 7, 14, 14, 14],
         {11: "0.000E+00", 12: "0.000E+00", 13: "0.000E+00"})

    # ── 04 도메인별통계 ──────────────────────────────────────────────────────
    ws = sheet(wb, "04_도메인별통계", "도메인별 집계", "도메인은 물성의 물리 분류다.")
    head(ws, 4, ["도메인", "정의 수", "값 있는 정의", "물성값 수", "보유 재료수", "실측(t1) 비율", "가정(t4) 비율"])
    rows = []
    for dom, nd in c.execute("select domain, count(*) from property_definition group by 1 order by 2 desc"):
        used = q1("""select count(distinct pd.key) from property_definition pd
                     join property_value pv on pv.property_key=pd.key where pd.domain=?""", dom)
        npv = q1("""select count(*) from property_value pv join property_definition pd
                    on pd.key=pv.property_key where pd.domain=?""", dom)
        nmat = q1("""select count(distinct pv.material_id) from property_value pv
                     join property_definition pd on pd.key=pv.property_key where pd.domain=?""", dom)
        t1 = q1("""select count(*) from property_value pv join property_definition pd
                   on pd.key=pv.property_key where pd.domain=? and pv.quality_tier=1""", dom)
        t4 = q1("""select count(*) from property_value pv join property_definition pd
                   on pd.key=pv.property_key where pd.domain=? and pv.quality_tier=4""", dom)
        rows.append((dom, nd, used, npv, nmat, t1 / npv if npv else 0, t4 / npv if npv else 0))
    body(ws, 5, rows, [16, 10, 12, 12, 12, 13, 13], {6: "0.0%", 7: "0.0%"})

    # ── 05 재료목록 ──────────────────────────────────────────────────────────
    ws = sheet(wb, "05_재료목록", "재료 전체 목록과 보유 물성",
               "해석 가능 열은 그 해석의 필수 물성이 전부 있는지를 뜻한다. 값의 정확도가 아니라 '입력이 갖춰졌는가'다.")
    head(ws, 4, ["ID", "재료명", "분류", "물성값 수", "물성 종류 수", "실측(t1)", "가정(t4)",
                 "출처 수", "구조", "낙하", "열전달", "열응력", "벤딩"])
    own = defaultdict(set)
    for mid, k in c.execute("select material_id, property_key from property_value"):
        own[mid].add(k)
    E, NU, RHO = "mechanical.youngs_modulus", "mechanical.poisson_ratio", "physical.density"
    PLAS = ("mechanical.yield_strength", "mechanical.tensile_strength", "mechanical.elongation_at_break")
    RATE = ("mechanical.cowper_symonds_c", "mechanical.dynamic_increase_factor",
            "mechanical.yield_strength_at_rate", "mechanical.johnson_cook_c")
    PRONY = ("mechanical.prony_relaxation_time", "mechanical.prony_shear_modulus",
             "mechanical.prony_tensile_modulus", "mechanical.prony_relative_modulus", "mechanical.shear_modulus")
    yes = lambda b: "O" if b else ""
    rows = []
    for mid, nm, cat in c.execute("select id, name, category from material order by category, name"):
        ks = own.get(mid, set())
        tt = dict(c.execute("select quality_tier, count(*) from property_value where material_id=? group by 1", (mid,)).fetchall())
        nsrc = q1("select count(distinct source_id) from property_value where material_id=?", mid)
        base = all(k in ks for k in (E, NU, RHO))
        rows.append((mid, nm, cat, sum(tt.values()), len(ks), tt.get(1, 0), tt.get(4, 0), nsrc,
                     yes(base and any(k in ks for k in PLAS)),
                     yes(base and all(k in ks for k in PLAS) and any(k in ks for k in RATE)),
                     yes(all(k in ks for k in ("thermal.conductivity", "thermal.specific_heat", RHO))),
                     yes(all(k in ks for k in ("thermal.expansion_linear", E, NU, RHO))),
                     yes(base and any(k in ks for k in PRONY))))
    body(ws, 5, rows, [6, 52, 11, 10, 12, 9, 9, 8, 7, 7, 8, 8, 7])

    # ── 06 해석준비도 ────────────────────────────────────────────────────────
    ws = sheet(wb, "06_해석준비도", "해석 종류별 준비도",
               "필수 = 그 물성이 하나라도 없으면 카드를 못 만드는 것. 선택군 = 여럿 중 하나만 있으면 되는 것.")
    ANALYSES = [
        ("구조 (정적)", "*MAT_024 / *MAT_001", [E, NU, RHO], [PLAS], None),
        ("낙하 (고속 충격)", "*MAT_024 + LCSR / *MAT_098", [E, NU, RHO] + list(PLAS), [RATE], None),
        ("열전달", "*MAT_THERMAL_ISOTROPIC", ["thermal.conductivity", "thermal.specific_heat", RHO], [], None),
        ("열응력", "*MAT_ADD_THERMAL_EXPANSION", ["thermal.expansion_linear", E, NU, RHO], [], None),
        ("벤딩 (폴더블)", "*MAT_006 / Prony", [E, NU, RHO], [PRONY], "film"),
        ("결로 (표면)", "열해석 + 젖음성", ["thermal.conductivity", "thermal.specific_heat", RHO],
         [("physical.contact_angle_water", "physical.surface_energy")], None),
        ("결로 (투습)", "열해석 + 습기확산", ["thermal.conductivity", "thermal.specific_heat", RHO],
         [("physical.water_vapor_transmission", "physical.gas_permeability_h2o",
           "physical.diffusion_coefficient", "chemical.water_absorption_24h")], "absorbent"),
    ]
    FILMISH = ("film", "tape", "oca", "ocr", "psa", "adhesive", "coating", "laminate",
               "foil", "sheet", "pi base", "foam")
    mats = {i: (n, cat) for i, n, cat in c.execute("select id,name,category from material")}
    head(ws, 4, ["해석", "LS-DYNA 카드", "대상 재료수", "필수 완비", "완비율", "선택군까지", "최종 비율", "최대 병목"])
    rows = []
    for label, card, must, anyof, filt in ANALYSES:
        if filt == "film":
            tgt = [m for m in mats if any(w in mats[m][0].lower() for w in FILMISH)]
        elif filt == "absorbent":
            tgt = [m for m in mats if mats[m][1] in ("polymer", "composite", "rubber", "foam")]
        else:
            tgt = list(mats)
        full = [m for m in tgt if all(k in own.get(m, set()) for k in must)]
        ready = [m for m in full if all(any(k in own.get(m, set()) for k in g) for g in anyof)]
        miss = sorted(((k, sum(1 for m in tgt if k not in own.get(m, set()))) for k in must), key=lambda x: -x[1])
        bott = f"{miss[0][0].split('.',1)[1]} 없음 {miss[0][1]}종" if miss and miss[0][1] else "필수는 모두 보유"
        if anyof:
            gm = sum(1 for m in full if not all(any(k in own.get(m, set()) for k in g) for g in anyof))
            if gm > (miss[0][1] if miss else 0):
                bott = f"선택군 전무 {gm}종"
        rows.append((label, card, len(tgt), len(full), len(full) / len(tgt),
                     len(ready) if anyof else "", (len(ready) / len(tgt)) if anyof else "", bott))
    body(ws, 5, rows, [20, 28, 12, 11, 10, 12, 10, 34], {5: "0%", 7: "0%"})

    r0 = 5 + len(rows) + 2
    ws.cell(row=r0, column=1, value="교차 병목 — 여러 해석을 동시에 막는 물성").font = TITLE_FONT
    head(ws, r0 + 1, ["물성", "결측 재료수", "막는 해석"])
    cross = [(E, "구조·낙하·벤딩·열응력"), (NU, "구조·낙하·벤딩·열응력"), (RHO, "전 동해석"),
             ("mechanical.yield_strength", "구조·낙하"), ("thermal.conductivity", "열전달·결로"),
             ("thermal.specific_heat", "열전달·결로"), ("thermal.expansion_linear", "열응력")]
    body(ws, r0 + 2, [(k.split(".", 1)[1], sum(1 for m in mats if k not in own.get(m, set())), why)
                      for k, why in cross], [30, 12, 30])

    # ── 07 재료물성매트릭스 ──────────────────────────────────────────────────
    ws = sheet(wb, "07_커버리지매트릭스", "분류 × 주요 물성 커버리지",
               "각 칸은 그 분류의 재료 중 해당 물성을 가진 비율이다.")
    KEYS = [E, NU, RHO, "mechanical.yield_strength", "mechanical.tensile_strength",
            "mechanical.elongation_at_break", "thermal.conductivity", "thermal.specific_heat",
            "thermal.expansion_linear", "thermal.glass_transition", "electrical.dielectric_constant",
            "electrical.resistivity_volume", "optical.refractive_index", "interface.peel_strength"]
    cats = [r[0] for r in c.execute("select category, count(*) n from material group by 1 order by n desc")]
    head(ws, 4, ["물성"] + [f"{x}\n({sum(1 for m in mats if mats[m][1]==x)}종)" for x in cats] + ["전체"])
    rows = []
    for k in KEYS:
        row = [k.split(".", 1)[1]]
        for cat in cats:
            tgt = [m for m in mats if mats[m][1] == cat]
            row.append(sum(1 for m in tgt if k in own.get(m, set())) / len(tgt) if tgt else 0)
        row.append(sum(1 for m in mats if k in own.get(m, set())) / len(mats))
        rows.append(tuple(row))
    body(ws, 5, rows, [30] + [11] * (len(cats) + 1), {j: "0%" for j in range(2, len(cats) + 3)})

    # ── 08 데이터품질 ────────────────────────────────────────────────────────
    ws = sheet(wb, "08_데이터품질", "데이터 품질 점검",
               "정합성 검사 23항목(integrity_check.py)과 같은 기준이다. 이상은 0이어야 한다.")
    head(ws, 4, ["점검 항목", "건수", "판정", "의미"])
    CHECKS = [
        ("출처 없는 값", "select count(*) from property_value where source_id is null", "근거 없는 값은 저장하지 않는다"),
        ("정의 없는 물성키", """select count(*) from property_value pv left join property_definition pd
            on pd.key=pv.property_key where pd.key is null""", "taxonomy 밖의 키"),
        ("단위가 정의와 다름", """select count(*) from property_value pv join property_definition pd
            on pd.key=pv.property_key where pv.value_num is not null and pd.si_unit is not null
            and replace(replace(replace(coalesce(pv.unit,''),'(',''),')',''),' ','')
             <> replace(replace(replace(pd.si_unit,'(',''),')',''),' ','')""", "단위 미환산·오입력"),
        ("조건이 dict 아님", "select count(*) from property_value where conditions is not null and conditions not like '{%'",
         "조건이 구조화되지 않으면 조회가 빗나간다"),
        ("가정값인데 tier4·estimated 아님", """select count(*) from property_value
            where instr(coalesce(conditions,''),'assumption')>0 and (quality_tier<>4 or method<>'estimated')""",
         "가정이 실측을 이기면 안 된다"),
        ("파장 없는 광학값(tier<4)", """select count(*) from property_value where property_key in
            ('optical.refractive_index','optical.reflectance','optical.extinction_coefficient','optical.birefringence')
            and quality_tier<4 and (conditions is null or (conditions not like '%wavelength%' and conditions not like '%line%'))""",
         "파장은 광학 물성의 정의 조건"),
        ("온도가 값인데 온도 조건", """select count(*) from property_value where property_key in
            ('thermal.melting_point','thermal.glass_transition','thermal.max_service_temp')
            and conditions like '%temperature_C%'""", "융점에 측정온도는 성립하지 않는다"),
        ("Prony 항에 항번호 없음", """select count(*) from property_value where property_key like 'mechanical.prony_%'
            and (conditions is null or conditions not like '%term%')""", "항번호 없으면 세트 복원 불가"),
        ("journal인데 DOI 없음", "select count(*) from source where kind='journal' and coalesce(doi,'')=''",
         "논문은 DOI로 재확인 가능해야 한다"),
        ("물성값 0인 재료", "select count(*) from material m where not exists(select 1 from property_value where material_id=m.id)",
         "빈 재료"),
    ]
    body(ws, 5, [(lbl, q1(sql), "OK" if q1(sql) == 0 else "확인 필요", why) for lbl, sql, why in CHECKS],
         [34, 9, 12, 44])

    r0 = 5 + len(CHECKS) + 2
    ws.cell(row=r0, column=1, value="조건(conditions) 보유율 — 조건이 필수인 물성").font = TITLE_FONT
    head(ws, r0 + 1, ["물성", "전체", "조건 보유", "보유율"])
    COND_NEED = ["thermal.conductivity", "optical.refractive_index", "mechanical.shear_modulus",
                 "mechanical.yield_strength_at_rate", "interface.peel_strength",
                 "electrical.dielectric_constant", "mechanical.creep_rate"]
    body(ws, r0 + 2, [(k.split(".", 1)[1],
                       q1("select count(*) from property_value where property_key=?", k),
                       q1("select count(*) from property_value where property_key=? and conditions is not null", k),
                       (q1("select count(*) from property_value where property_key=? and conditions is not null", k)
                        / max(q1("select count(*) from property_value where property_key=?", k), 1)))
                      for k in COND_NEED], [30, 10, 12, 10], {4: "0%"})

    # ── 09 전체물성값 ────────────────────────────────────────────────────────
    ws = sheet(wb, "09_전체물성값", "물성값 전체 목록 (원본)",
               "모든 값이 SI 단위로 정규화돼 있다. 조건이 다르면 같은 물성도 별개 행이다 — 그게 이 표가 긴 이유다.")
    head(ws, 4, ["재료ID", "재료명", "분류", "물성 키", "물성명", "값", "값(텍스트)", "단위",
                 "등급", "방법", "조건", "출처", "DOI/URL"])
    rows = c.execute("""select pv.material_id, m.name, m.category, pv.property_key, coalesce(pd.name,''),
            pv.value_num, coalesce(pv.value_text,''), coalesce(pv.unit,''), pv.quality_tier, pv.method,
            coalesce(pv.conditions,''), substr(coalesce(s.title,''),1,80),
            coalesce(nullif(s.doi,''), s.url, '')
        from property_value pv join material m on m.id=pv.material_id
        left join property_definition pd on pd.key=pv.property_key
        left join source s on s.id=pv.source_id
        order by m.category, m.name, pv.property_key, pv.quality_tier, pv.id""").fetchall()
    body(ws, 5, rows, [7, 40, 10, 34, 22, 14, 18, 14, 6, 11, 44, 52, 34], {6: "0.0000E+00"})
    ws.auto_filter.ref = f"A4:M{4 + len(rows)}"

    # ── 10 해석별필요물성 ────────────────────────────────────────────────────
    ws = sheet(wb, "10_해석별필요물성", "해석별로 무엇이 필요한가 (참조)",
               "'필수'가 하나라도 없으면 카드를 만들 수 없다. '선택'은 있으면 정확도가 올라간다.")
    head(ws, 4, ["해석", "구분", "물성", "SI 단위", "왜 필요한가", "현재 보유 재료수"])
    NEED = [
        ("구조 (정적)", "필수", E, "탄성 강성"), ("구조 (정적)", "필수", NU, "횡변형·체적거동"),
        ("구조 (정적)", "필수", RHO, "질량·관성"),
        ("구조 (정적)", "필수", "mechanical.yield_strength", "소성 시작점"),
        ("구조 (정적)", "선택", "mechanical.tensile_strength", "경화 기울기 ETAN 산출"),
        ("구조 (정적)", "선택", "mechanical.elongation_at_break", "파단 판정 FAIL"),
        ("낙하 (고속 충격)", "필수", "mechanical.yield_strength_at_rate", "LCSR 배율 곡선의 원자료"),
        ("낙하 (고속 충격)", "선택", "mechanical.cowper_symonds_c", "2상수 율속 근사(C)"),
        ("낙하 (고속 충격)", "선택", "mechanical.cowper_symonds_p", "2상수 율속 근사(p)"),
        ("낙하 (고속 충격)", "선택", "mechanical.johnson_cook_c", "JC 율속감도 — 없으면 JC를 써도 율속 효과가 없다"),
        ("낙하 (고속 충격)", "선택", "mechanical.johnson_cook_m", "온도연화. 솔더는 T/Tm≈0.6이라 중요"),
        ("열전달", "필수", "thermal.conductivity", "전도 계산"),
        ("열전달", "필수", "thermal.specific_heat", "축열·과도 응답"),
        ("열전달", "선택", "optical.emissivity_total", "복사 경계조건"),
        ("열응력", "필수", "thermal.expansion_linear", "열변형 구동항"),
        ("열응력", "선택", "thermal.glass_transition", "Tg 위아래로 CTE가 바뀐다"),
        ("벤딩 (폴더블)", "필수", "mechanical.prony_relative_modulus", "완화 스펙트럼 계수"),
        ("벤딩 (폴더블)", "필수", "mechanical.prony_relaxation_time", "완화 시간(계수와 짝)"),
        ("벤딩 (폴더블)", "선택", "mechanical.shear_modulus", "Prony가 없을 때 차선"),
        ("벤딩 (폴더블)", "선택", "structure.layer_thickness", "중립축 위치"),
        ("결로", "필수", "physical.contact_angle_water", "표면 응축 젖음 거동"),
        ("결로", "선택", "physical.water_vapor_transmission", "투습 경로"),
        ("결로", "선택", "physical.diffusion_coefficient", "수분 확산"),
        ("초탄성 (고무·폼)", "필수", "mechanical.hyperelastic_coefficient", "Ogden·Mooney-Rivlin 계수"),
        ("초탄성 (고무·폼)", "필수", "mechanical.hyperelastic_exponent", "같은 모델의 지수(계수와 짝)"),
    ]
    units = dict(c.execute("select key, si_unit from property_definition").fetchall())
    body(ws, 5, [(a, b, k.split(".", 1)[1], units.get(k, ""), why,
                  q1("select count(distinct material_id) from property_value where property_key=?", k))
                 for a, b, k, why in NEED], [18, 7, 34, 14, 52, 14])

    # ── 11 해석 커버리지 — docx와 같은 숫자를 보게 한다 ────────────────────
    import sys as _sys
    _sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
    from coverage_report import ANALYSES as _AN
    from coverage_report import compute as _cov
    _c, _mat, _own, cov = _cov()
    ws = sheet(wb, "11_해석커버리지",
               "해석별 물성 커버리지 — 셀 채움률과 재료 준비율",
               "셀 채움률 = 채워진 (재료 x 필수물성) 칸 / 전체 칸 (수집 진척도). "
               "재료 준비율 = 그 해석을 실제로 돌릴 수 있는 재료 / 대상 재료 (해석 가능성). "
               "둘은 다른 것을 잰다 — 재료마다 비어 있는 칸이 다르면 셀은 높고 준비는 낮다.")
    head(ws, 4, ["해석", "설명", "대상 재료", "전체 칸", "채워진 칸", "셀 채움률",
                 "준비 재료", "재료 준비율", "가장 큰 공백", "미보유 재료"])
    rows = []
    for x in cov:
        desc = next(a[4] for a in _AN if a[0] == x["name"])
        top = x["missing"][0] if x["missing"] else ("-", 0)
        rows.append((x["name"], desc, x["n_target"], x["cells"], x["filled"],
                     x["cell_pct"] / 100, x["n_ready"], x["ready_pct"] / 100,
                     top[0].replace("택일군: ", "").split(".", 1)[-1], top[1]))
    body(ws, 5, rows, [20, 46, 11, 10, 11, 12, 11, 13, 44, 12])
    for r in range(5, 5 + len(rows)):
        for col in ("F", "H"):
            ws[f"{col}{r}"].number_format = "0.0%"
    tot_c = sum(x["cells"] for x in cov)
    tot_f = sum(x["filled"] for x in cov)
    n = 5 + len(rows) + 1
    ws[f"A{n}"] = "전체"
    ws[f"A{n}"].font = Font(bold=True)
    ws[f"D{n}"], ws[f"E{n}"] = tot_c, tot_f
    ws[f"F{n}"] = tot_f / tot_c
    ws[f"F{n}"].number_format = "0.0%"
    ws[f"F{n}"].font = Font(bold=True)

    wb.save(OUT)
    print(f"저장: {OUT}")
    print(f"  시트 {len(wb.sheetnames)}개 — {', '.join(wb.sheetnames)}")


if __name__ == "__main__":
    main()
